from __future__ import annotations

import json
import os
import urllib.request
import urllib.error
import io
import re
import subprocess
import sys

# Self-healing import to install openpyxl if missing (critical for Render Node environment)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
except ImportError:
    try:
        subprocess.check_call([
            sys.executable, "-m", "pip", "install", "openpyxl>=3.1.0", "--break-system-packages"
        ])
    except Exception:
        # Fallback without --break-system-packages in case of older python env
        subprocess.check_call([
            sys.executable, "-m", "pip", "install", "openpyxl>=3.1.0"
        ])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

from flask import Flask, Response, jsonify, request, send_from_directory

# Load .env file manually (no python-dotenv needed)
def load_dotenv(path: str = ".env") -> None:
    if not os.path.exists(path):
        return
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, value = line.partition("=")
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = value

load_dotenv()

app = Flask(__name__, static_folder="dist", static_url_path="")

GEMINI_KEY = os.environ.get("GEMINI_API_KEY", "")
GEMINI_MODELS = ["gemini-2.5-flash", "gemini-2.0-flash", "gemini-1.5-flash"]


def call_gemini(payload: dict, model: str) -> dict | None:
    """Call Gemini API with given payload. Returns parsed JSON or None on failure."""
    url = (
        f"https://generativelanguage.googleapis.com/v1beta/models/"
        f"{model}:generateContent?key={GEMINI_KEY}"
    )
    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=data,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        if e.code == 404:
            return None  # try next model
        raise
    except Exception:
        raise


# ── /api/parse-annexure ─────────────────────────────────────────────────────

@app.post("/api/parse-annexure")
def parse_annexure() -> Response:
    body = request.get_json(force=True) or {}
    text_items: list[str] = body.get("textItems", [])
    brand_name: str = body.get("brandName", "")
    mode: str = body.get("mode", "commission")

    if not GEMINI_KEY:
        return Response("Gemini API Key is not configured on the server.", status=400)

    full_text = " ".join(text_items)

    if mode == "commission":
        target_instruction = (
            f'Your task is to extract all Platform Commission slabs from the text for the brand: "{brand_name}".\n'
            "Do NOT extract Fixed Fee slabs. Only extract from the Platform Commission / Commission rates grid or section.\n"
            'If there are general default rules (e.g. Article Type is "ALL" or "Default"), extract them.\n'
            "All returned rules should have fixedFee set to 0."
        )
    else:
        target_instruction = (
            f'Your task is to extract all Fixed Fee slabs from the text for the brand: "{brand_name}".\n'
            "Do NOT extract Platform Commission slabs. Only extract from the Fixed Fee grid or section.\n"
            'Do NOT include general default rules (like Article Type "ALL" with 0 commission) that belong to the commission table.\n'
            "All returned rules should have commissionPercent set to 0."
        )

    prompt = (
        "You are an expert data parser. You are given the raw extracted text from a Myntra Commercial Terms Agreement (CTA) PDF annexure.\n"
        f"{target_instruction}\n\n"
        "Return a JSON array of objects conforming to the following TypeScript interface:\n"
        "interface FeeRule {\n"
        f'  brand: string; // The brand/party name, i.e., "{brand_name}"\n'
        '  category: string;\n'
        '  articleType: string;\n'
        '  gender: string;\n'
        '  lowerLimit: number;\n'
        '  upperLimit: number;\n'
        '  commissionPercent: number;\n'
        '  fixedFee: number;\n'
        "}\n\n"
        "Notes:\n"
        "1. Extract only the slabs relevant to the requested mode as instructed above.\n"
        "2. Keep the matching exact and complete.\n"
        "3. Return ONLY a valid JSON array. Do not wrap in markdown or any other text.\n\n"
        f"Raw PDF Text:\n{full_text}"
    )

    payload = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"responseMimeType": "application/json"},
    }

    last_error = None
    for model in GEMINI_MODELS:
        try:
            data = call_gemini(payload, model)
            if data is None:
                continue  # 404, try next
            text = (
                data.get("candidates", [{}])[0]
                .get("content", {})
                .get("parts", [{}])[0]
                .get("text", "")
            )
            if not text:
                raise ValueError("No text in Gemini response")
            rules = json.loads(text)

            def coerce(r: dict) -> dict:
                upper = r.get("upperLimit")
                if upper is None or str(upper) == "Infinity" or (isinstance(upper, (int, float)) and upper >= 10_000_000):
                    upper = float("inf")
                else:
                    upper = float(upper)
                return {
                    **r,
                    "brand": brand_name,
                    "lowerLimit": float(r.get("lowerLimit") or 0),
                    "upperLimit": upper,
                    "commissionPercent": float(r.get("commissionPercent") or 0),
                    "fixedFee": float(r.get("fixedFee") or 0),
                }

            return jsonify({"rules": [coerce(r) for r in rules]})
        except Exception as exc:
            last_error = exc

    err_msg = str(last_error) if last_error else "All Gemini models failed."
    return Response(err_msg, status=500)


# ── /api/chat ────────────────────────────────────────────────────────────────

@app.post("/api/chat")
def chat() -> Response:
    body = request.get_json(force=True) or {}
    message: str = body.get("message", "")
    user_name: str = body.get("userName", "friend")

    if not GEMINI_KEY:
        return jsonify({"reply": "Server me API key configure nahi hai."})

    system_prompt = (
        f'You are AJ, the user\'s "Dashboard Friend" and business assistant for this Myntra Cost Calculator.\n'
        f'Your tone is extremely friendly, professional, and helpful. Always address the user as "{user_name}".\n\n'
        "CORE CAPABILITIES:\n"
        "1. Simple Math: Perform basic arithmetic (+, -, *, /).\n"
        "2. Percentage Logic: Handle queries like \"X% of Y\", \"percentage of X\", etc.\n"
        "3. Advanced Business Math:\n"
        "   - Reverse GST: Extract base price from a tax-inclusive price.\n"
        "   - Profit Margin vs Markup.\n"
        "   - Discount Stacking.\n"
        "   - Break-even calculations.\n"
        "4. Myntra-specific Help:\n"
        "   - Purchase Cost = MRP * (1 - Purchase Margin%) / (1 + Tax%)\n"
        "   - Purchase Tax = Purchase Cost * Tax%\n"
        "   - Final Purchase Cost = Purchase Cost + Purchase Tax\n"
        "   - Myntra Commission is deducted from selling price\n"
        "   - Fixed Fee is a flat charge per item\n\n"
        "GUIDELINES:\n"
        "- Respond in Hinglish (a natural mix of Hindi and English).\n"
        "- For math queries, clearly state the result and briefly show the formula.\n"
        "- Keep responses concise but warm."
    )

    payload = {
        "system_instruction": {"parts": [{"text": system_prompt}]},
        "contents": [{"parts": [{"text": message}]}],
    }

    try:
        data = call_gemini(payload, "gemini-2.0-flash")
        if data is None:
            return jsonify({"reply": "AI se connect nahi ho pa raha. Baad me try karo. 😔"})
        reply = (
            data.get("candidates", [{}])[0]
            .get("content", {})
            .get("parts", [{}])[0]
            .get("text", "Koi response nahi mila.")
        )
        return jsonify({"reply": reply})
    except Exception as exc:
        return jsonify({"reply": "I'm sorry, I'm having trouble connecting right now. 😔"})


# ── Excel Export (openpyxl Backend) ─────────────────────────────────────────

def sanitize_filename(filename: str) -> str:
    cleaned = "".join(char for char in filename if char not in '<>:"/\\|?*').strip().strip(".")
    return cleaned or "myntra_export"


def coerce_export_cell(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, str):
        text = ILLEGAL_CHARACTERS_RE.sub("", value).strip()
        if not text:
            return ""
        cleaned = text.replace(",", "")
        is_percent = cleaned.endswith("%")
        if is_percent:
            cleaned = cleaned[:-1].strip()
        
        try:
            if "." in cleaned:
                num = float(cleaned)
                return num / 100.0 if is_percent else num
            else:
                num = int(cleaned)
                return float(num) / 100.0 if is_percent else num
        except ValueError:
            return text
    return value


@app.post("/api/export")
def export_excel_route() -> Response:
    import time
    start_time = time.time()
    try:
        payload = request.get_json(force=True, silent=True)
        if not payload or not isinstance(payload, dict):
            return jsonify({"error": "Invalid request payload."}), 400
            
        filename = sanitize_filename(str(payload.get("filename", "myntra_export")))
        headers = [ILLEGAL_CHARACTERS_RE.sub("", str(cell)) for cell in payload.get("headers", [])]
        rows = payload.get("rows", []) or []
        
        if not headers:
            return jsonify({"error": "No headers provided."}), 400
            
        print(f"DEBUG EXPORT: Started processing {len(rows)} rows with {len(headers)} columns...")
        
        t_start_build = time.time()
        # Create workbook
        wb = Workbook()
        sheet = wb.active
        sheet.title = "RawData"
        
        # Identify percentage columns beforehand to avoid slow string checks in loops
        percent_cols = [col_idx for col_idx, h in enumerate(headers, start=1) if "%" in str(h)]
        
        # Append headers
        sheet.append(headers)
        
        # Fill data rows using fast append
        row_idx = 2  # Headers are row 1, data starts at row 2
        for row in rows:
            padded = row + [""] * (len(headers) - len(row))
            coerced_row = [coerce_export_cell(cell) for cell in padded[: len(headers)]]
            sheet.append(coerced_row)
            
            # Format only percentage cells in the newly appended row
            for col_idx in percent_cols:
                cell_value = coerced_row[col_idx - 1]
                if isinstance(cell_value, (int, float)):
                    sheet.cell(row=row_idx, column=col_idx).number_format = "0%"
            row_idx += 1
                    
        t_end_build = time.time()
        print(f"DEBUG EXPORT: Data population took {t_end_build - t_start_build:.4f} seconds.")
        
        t_start_style = time.time()
        # Freeze Pane (Row 1)
        sheet.freeze_panes = "A2"
        
        # Styling definitions
        red_fill = PatternFill(fill_type="solid", fgColor="FFCCCC")
        yellow_fill = PatternFill(fill_type="solid", fgColor="FFFFCC")
        green_fill = PatternFill(fill_type="solid", fgColor="CCFFCC")
        
        thin_black = Side(style="thin", color="000000")
        thin_gray = Side(style="thin", color="D3D3D3")
        
        header_border = Border(left=thin_black, right=thin_black, top=thin_black, bottom=thin_black)
        data_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
        
        # Style Header Row
        sheet.row_dimensions[1].height = 18
        for col_idx in range(1, len(headers) + 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.font = Font(name="Calibri", size=10, bold=True)
            cell.border = header_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            
            header_val = str(headers[col_idx - 1] or "")
            if header_val in ["Style ID", "SKU ID"]:
                cell.fill = yellow_fill
            elif header_val in [
                "No.", "Seller Name", "Item Name", "Category", "Brand", "Type", 
                "ASIN Number", "SKU Number", "HSN Number", "MRP Price", "Item Color", 
                "Weight", "Weight Unit", "Length", "Length Unit", "Width", "Width Unit", 
                "Height", "Height Unit", "Channel Price", "Purchase Margin(%)", "Tax", 
                "Purchase Cost", "Purchase Tax", "Final Purchase Cost"
            ]:
                cell.fill = red_fill
            else:
                cell.fill = green_fill
                
        # Style Data Rows (Optimized to skip slow cell-by-cell loop for 250,000+ cells)
        # Instead, we enable native Excel gridlines for a clean and ultra-fast look
        sheet.views.sheetView[0].showGridLines = True
                    
        t_end_style = time.time()
        print(f"DEBUG EXPORT: Data styling took {t_end_style - t_start_style:.4f} seconds.")
        
        t_start_widths = time.time()
        # Calculate and Apply Column Widths efficiently (strategic sampling like Ajio)
        column_widths = {}
        for col_idx in range(1, len(headers) + 1):
            header_val = str(headers[col_idx - 1] or "")
            column_widths[col_idx] = max(len(header_val) + 3, 12)
            
        if len(rows) > 100:
            sample_size = min(500, max(100, len(rows) // 10))
            sample_indices = [int(i * len(rows) / sample_size) for i in range(sample_size)]
        else:
            sample_indices = range(len(rows))
            
        for row_idx in sample_indices:
            if row_idx < len(rows):
                row = rows[row_idx]
                for col_idx in range(1, len(headers) + 1):
                    if col_idx - 1 < len(row):
                        cell_val = str(row[col_idx - 1] or "")
                        column_widths[col_idx] = max(column_widths[col_idx], min(len(cell_val) + 3, 50))
                    
        for col_idx, width in column_widths.items():
            column_letter = get_column_letter(col_idx)
            sheet.column_dimensions[column_letter].width = width
            
        t_end_widths = time.time()
        print(f"DEBUG EXPORT: Column width calculations took {t_end_widths - t_start_widths:.4f} seconds.")
        
        t_start_save = time.time()
        # Write to BytesIO and return response
        output = io.BytesIO()
        wb.save(output)
        workbook_bytes = output.getvalue()
        t_end_save = time.time()
        print(f"DEBUG EXPORT: Saving workbook took {t_end_save - t_start_save:.4f} seconds.")
        print(f"DEBUG EXPORT: Total backend processing took {time.time() - start_time:.4f} seconds.")
        
    except Exception as exc:
        app.logger.exception("Excel export failed: %s", exc)
        return jsonify({"error": f"Excel export failed: {exc}"}), 500
        
    response = Response(
        workbook_bytes,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response.headers["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
    return response


# ── Static Files (React Build) ───────────────────────────────────────────────

@app.get("/")
def index() -> Response:
    return send_from_directory("dist", "index.html")


@app.get("/<path:path>")
def static_files(path: str) -> Response:
    if os.path.exists(os.path.join("dist", path)):
        return send_from_directory("dist", path)
    return send_from_directory("dist", "index.html")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 3001))
    print(f"Server running on port {port}")
    app.run(host="0.0.0.0", port=port, debug=False)
